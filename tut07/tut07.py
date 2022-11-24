from datetime import datetime
start_time = datetime.now()

def octant_analysis(mod):
    # adding the values for avarage
    
    # print(df['U']-u1_avg)
    #  finding the values of "V'=V - V avg" and insert in outfile using" db.at" function
    for i in range(0,n):
        df.at[i,"U'=U - U avg"]=df['U'][i]-u1_avg
    for i in range(0,n):
        df.at[i,"V'=V - V avg"]=df['V'][i]-v1_avg
    for i in range(0,n):
        df.at[i,"W'=W - W avg"]=df['W'][i]-w1_avg


    octant_list=[]
    rows = len(df)


    # finding the value of octant ans insert in output file
    try:
        for i in range(0,n):
            if(df.at[i,"U'=U - U avg"]>0 and df.at[i,"V'=V - V avg"]>0 and df.at[i,"W'=W - W avg"]>0):
                df.at[i,"Octant"]=1
                octant_list.append(df.at[i,'Octant'])
                
            if(df.at[i,"U'=U - U avg"]>0 and df.at[i,"V'=V - V avg"]>0 and df.at[i,"W'=W - W avg"]<0):
                df.at[i,'Octant']=-1
                octant_list.append(df.at[i,'Octant'])
                
            if(df.at[i,"U'=U - U avg"]<0 and df.at[i,"V'=V - V avg"]>0 and df.at[i,"W'=W - W avg"]>0):
                df.at[i,'Octant']=2
                octant_list.append(df.at[i,'Octant'])
                
            if(df.at[i,"U'=U - U avg"]<0 and df.at[i,"V'=V - V avg"]>0 and df.at[i,"W'=W - W avg"]<0):
                df.at[i,'Octant']=-2
                octant_list.append(df.at[i,'Octant'])
                
            if(df.at[i,"U'=U - U avg"]<0 and df.at[i,"V'=V - V avg"]<0 and df.at[i,"W'=W - W avg"]>0):
                df.at[i,'Octant']=3
                octant_list.append(df.at[i,'Octant'])
                
            if(df.at[i,"U'=U - U avg"]<0 and df.at[i,"V'=V - V avg"]<0 and df.at[i,"W'=W - W avg"]<0):
                    df.at[i,'Octant']=-3
                    octant_list.append(df.at[i,'Octant'])
                
            if(df.at[i,"U'=U - U avg"]>0 and df.at[i,"V'=V - V avg"]<0 and df.at[i,"W'=W - W avg"]>0):
                df.at[i,'Octant']=4
                octant_list.append(df.at[i,'Octant'])
                
            if(df.at[i,"U'=U - U avg"]>0 and df.at[i,"V'=V - V avg"]<0 and df.at[i,"W'=W - W avg"]<0):
                    df.at[i,'Octant']=-4
                    octant_list.append(df.at[i,'Octant'])
    except:
        print("Octant_list not found")   


    octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
    l=[]

    try:
        # Calculating the octant values
        for i in range(0, rows):
            if df.at[i,"U'=U - U avg"] >= 0 and  df.at[i,"V'=V - V avg"] >= 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 1
                else:
                  df.at[i, 'Octant'] = -1
            elif df.at[i,"U'=U - U avg"] < 0 and  df.at[i,"V'=V - V avg"] >= 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 2
                else:
                  df.at[i, 'Octant'] = -2
            elif df.at[i,"U'=U - U avg"] < 0 and  df.at[i,"V'=V - V avg"] < 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 3
                else:
                  df.at[i, 'Octant'] = -3
            elif df.at[i,"U'=U - U avg"] >= 0 and  df.at[i,"V'=V - V avg"] < 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 4
                else:
                  df.at[i, 'Octant'] = -4
            l.append(df.at[i, 'Octant'])

        df.at[1, ''] = "User Input"
        df.at[0, 'Octant ID'] = "Overall Count"
        df.at[1, 'Octant ID'] = "Mod "+ str(mod) 

        df.at[0, "1"] = l.count(1)
        df.at[0, "-1"] = l.count(-1)
        df.at[0 ,"2"] = l.count(2)
        df.at[0 ,"-2"] = l.count(-2)
        df.at[0 ,"3"] = l.count(3)
        df.at[0 ,"-3"] = l.count(-3)
        df.at[0 ,"4"] = l.count(4)
        df.at[0 ,"-4"] = l.count(-4)
    except:
        print("Error in counting octant values.")
        exit()

    try:
        # Split list into ranges and find the count of octant values
        start = 0
        end = len(l)
        step = int(mod)
        idx=2
        total_rows_mod = math.ceil(rows/step)
        for i in range(start, end, step):
            x = i
            sub_list = l[x:x+step]
            y = x+step-1
            if y>rows:
                y=rows-1
            df.at[idx ,'Octant ID'] = str(x)+"-"+str(y)
            df.at[idx, '1'] = sub_list.count(1)
            df.at[idx, '-1'] = sub_list.count(-1)
            df.at[idx, '2'] = sub_list.count(2)
            df.at[idx, '-2'] = sub_list.count(-2)
            df.at[idx, '3'] = sub_list.count(3)
            df.at[idx, '-3'] = sub_list.count(-3)
            df.at[idx, '4'] = sub_list.count(4)
            df.at[idx, '-4'] = sub_list.count(-4)
            idx+=1
    except:
        print("Error in counting octant values for ranges.")
        exit()
        
    try:
        ### Inserting Rank Columns 
        col_num = 21
        for i in range(1,5):
            df.insert(col_num, column="Rank of "+str(i), value="")
            col_num+=1
            df.insert(col_num, column="Rank of "+str(-1*i), value="")
            col_num+=1
        df.insert(col_num, column="Rank 1 Octant ID", value="")
        col_num+=1
        df.insert(col_num, column="Rank 1 Octant Name", value="")
        col_num+=1
        
        ### Calculating rank for Overall Octant Count
        dict={}
        l=[]
        for i in range(1,5):
            oct_cnt = df.at[0, str(i)]
            dict[oct_cnt] = str(i)
            l.append(oct_cnt)
            oct_cnt = df.at[0, str(-1*i)]
            dict[oct_cnt] = str(-1*i)
            l.append(oct_cnt)
        
        l.sort(reverse=True)
        rank = 1
        df.at[0, "Rank 1 Octant ID"] = dict[l[0]]
        df.at[0, "Rank 1 Octant Name"] = octant_name_id_mapping[dict[l[0]]]
        
        for i in l:
            oct_id = "Rank of "+dict[i]
            df.at[0, oct_id] = rank
            rank+=1
        
        ## Calculating rank for Mod Octant Count
        rank1=[]
        for idx in range(2, total_rows_mod+2): 
            dict={}
            l=[]
            for i in range(1,5):
                oct_cnt = df.at[idx, str(i)]
                dict[oct_cnt] = str(i)
                l.append(oct_cnt)
                oct_cnt = df.at[idx, str(-1*i)]
                dict[oct_cnt] = str(-1*i)
                l.append(oct_cnt)

            l.sort(reverse=True)
            df.at[idx, "Rank 1 Octant ID"] = dict[l[0]]
            rank1.append(dict[l[0]])
            df.at[idx, "Rank 1 Octant Name"] = octant_name_id_mapping[dict[l[0]]]
            
            rank = 1
            for i in l:
                oct_id = "Rank of "+dict[i]
                df.at[idx, oct_id] = rank
                rank+=1
        
        ### Count of Rank 1 Mod Values
        idx = total_rows_mod+5
        df.at[idx, '1'] = "Octant ID"
        df.at[idx, '-1'] = "Octant Name"
        df.at[idx, '2'] = "Count of Rank 1 Mod Values"
        idx += 1
        for i in range(1,5):
            oct_id = str(i)
            cnt = rank1.count(oct_id)
            df.at[idx, '1'] = oct_id
            df.at[idx, '-1'] = octant_name_id_mapping[oct_id]
            df.at[idx, '2'] = cnt
            idx+=1
            
            oct_id = str(-1*i)
            cnt = rank1.count(oct_id)
            df.at[idx, '1'] = oct_id
            df.at[idx, '-1'] = octant_name_id_mapping[oct_id]
            df.at[idx, '2'] = cnt
            idx+=1
            
    except Exception as e:
        print("Error in calculating rank.", e)



    df2=pd.DataFrame(columns=['1','-1','2','-2','3','-3','4','-4'],index=['1','-1','2','-2','3','-3','4','-4'])
    df2=df2.fillna(0)


    df.at[0,'x']=''
    df.at[0,'Count']=''
    df.at[0,'Longest Subsquence Length']=''
    df.at[0,' Frequency']=''
    df.at[0, 'y']=''
    df.at[1,'Count']='1'
    df.at[2,'Count']='-1'
    df.at[3,'Count']='2'
    df.at[4,'Count']='-2'
    df.at[5,'Count']='3'
    df.at[6,'Count']='-3'
    df.at[7,'Count']='4'
    df.at[8,'Count']='-4'
    list_oct=[1,-1,2,-2,3,-3,4,-4]

    idx=1
    ptr = 0
    for i in list_oct:
        df.at[ptr, 'Octant ##'] = i
        ptr_tmp = ptr
        ptr += 1
        a=i
        l=0;
        r=0;
        ans=0
        while((l<=r&r<n)&(l<n)):    
            if((octant_list[l]==octant_list[r])&((octant_list[l]==a)&(octant_list[r]==a))):
                r=r+1
            else :
                # if(ans<r-l+1)
                ans=max(ans,r-l)
                l=r
                r=r+1
            if(r==n):
                ans=max(ans,r-l)
        
        
        l=0;
        r=0;
        ans1=0;
        cnt=0
        while((l<=r&r<n)&(l<n)):    
            if((octant_list[l]==octant_list[r])&((octant_list[l]==a)&(octant_list[r]==a))):
                r=r+1
            else :
                # if(ans<r-l+1)
                ans1=max(ans1,r-l)
                if(ans1==ans):
                    cnt=cnt+1
                    ans1=0
                l=r
                r=r+1
            if(r==n):
                ans1=max(ans1,r-l) 
                if(ans1==ans):
                    cnt=cnt+1
        count_tmp = 0
        strt = -1
        for numb in range(len(df['Octant'])):
            if df.at[numb, 'Octant'] == i:
                if strt == -1:
                    strt = numb
                count_tmp += 1
            else:
                count_tmp = 0
                strt = -1
            if count_tmp == ans:
                count_tmp = 0
                df.at[ptr, 'Count##'] = df.at[strt, 'T']
                df.at[ptr, 'Longest Sub Length'] = df.at[numb, 'T']
                ptr += 1
        df.at[idx,'Longest Subsquence Length']=ans
        df.at[ptr_tmp, 'Longest Sub Length'] = ans
        df.at[idx,' Frequency']=cnt
        df.at[ptr_tmp, 'Count##'] = cnt
        idx+=1

    


    df.at[0, 'AJ'] = ''
    i = 0
    for j in range(0,n-1):
        ro=str(int(octant_list[j]))
        co=str(int(octant_list[j+1]))
        df2.loc[ro,co]+=1
    df.at[i,'AK']="Overall Transition Count"
    df.at[i+2,'AK']="Count"
    i=i+1;
    df.at[i,'AL']="To"
    df.at[i+2,'AJ'] ='From'
    i=i+1
    df.at[i,'AL']='1'
    df.at[i,'AM']='-1'
    df.at[i,'AN']='2'
    df.at[i,'AO']='-2'
    df.at[i,'AP']='3'
    df.at[i,'AQ']='-3'
    df.at[i,'AR']='4'
    df.at[i,'AS']='-4'
    df.at[i+1,'AK']='1'
    df.at[i+2,'AK']='-1'
    df.at[i+3,'AK']='2'
    df.at[i+4,'AK']='-2'
    df.at[i+5,'AK']='3'
    df.at[i+6,'AK']='-3'
    df.at[i+7,'AK']='4'
    df.at[i+8,'AK']='-4'
    i=i+1
    col_index=41
    for k in range(0,8):
        for l in range(0,8):
            df.iloc[i+k,col_index+l]=df2.iloc[k,l]

    i=i+7
    st=0
    # lis2=np.array_split(octant_list, mod)     
    lis2 = [octant_list[i * mod:(i + 1) * mod] for i in range((len(octant_list) + mod - 1) // mod )] 
    for x in lis2:
        df2=pd.DataFrame(columns=['1','-1','2','-2','3','-3','4','-4'],index=['1','-1','2','-2','3','-3','4','-4'])
        df2=df2.fillna(0)
        # print(df2)

        # print(df2['1']['1'])
        print(len(x),end='  ')
        for j in range(0,len(x)-1):
            ro=str(int(x[j]))
            co=str(int(x[j+1]))
            df2.loc[ro,co]+=1
        
        i=i+3  
        df.at[i,'AK']="Mod Transition Count"
        df.at[i+2,'AK']="Count"
        if st+mod-1 >= len(octant_list):
            df.at[i+1,'AK']=str(st)+"-"+str(len(octant_list)-1)
        else:
            df.at[i+1,'AK']=str(st)+"-"+str(st+mod-1)
            first = int(df.at[st+mod-1, 'Octant'])
            second = int(df.at[st+mod, 'Octant'])
            df2.loc[str(first), str(second)]+=1
            ## print(first, second)

        st=st+mod
        
        i=i+1;
        df.at[i,'AL']="To"
        df.at[i+2,'AJ'] ='From'
        i=i+1
        df.at[i,'AL']='1'
        df.at[i,'AM']='-1'
        df.at[i,'AN']='2'
        df.at[i,'AO']='-2'
        df.at[i,'AP']='3'
        df.at[i,'AQ']='-3'
        df.at[i,'AR']='4'
        df.at[i,'AS']='-4'
        df.at[i+1,'AK']='1'
        df.at[i+2,'AK']='-1'
        df.at[i+3,'AK']='2'
        df.at[i+4,'AK']='-2'
        df.at[i+5,'AK']='3'
        df.at[i+6,'AK']='-3'
        df.at[i+7,'AK']='4'
        df.at[i+8,'AK']='-4'
        i=i+1
        col_index=41
        for k in range(0,8):
            for l in range(0,8):
                df.iloc[i+k,col_index+l]=df2.iloc[k,l]
        # print(df2)
        # print(df2.iloc[1][0])

        i=i+7

from platform import python_version
import pandas as pd
import os
import math
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, colors, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


for filename in os.listdir('input'):
    f = os.path.join('input', filename)
    if os.path.isfile(f) and f.endswith('.xlsx'):
        df = pd.read_excel(f)  


        mod=5000

        octant_analysis(mod)
        df.to_excel(os.path.join('output', filename), index=None)

        file = load_workbook('output.xlsx')
        sheet = file.active

        side = Side(border_style='thin', color="000000")
        border = Border(top=side, bottom=side, left=side, right=side)
        for cell in sheet['AG1':'AI10']:
            for x in cell:
                x.border = border
        for cell in sheet['AK1':'AM' + str(sum([sheet['AI' + str(i)].value for i in range(3,11)]) + 9)]:
            for x in cell:
                x.border = border
        for cell in sheet['M1':'AE' + str((len(df) // mod) + 4)]:
            for x in cell:
                x.border = border
        for cell in sheet['N' + str((len(df) // mod) + 8):'P' + str((len(df) // mod) + 16)]:
            for x in cell:
                x.border = border
        for i in range((len(df) // mod) + 2):
            for cell in sheet['AO' + str(4 + 13*i) : 'AW' + str(12 + 13*i)]:
                for x in cell:
                    x.border = border

        style =DifferentialStyle(fill=PatternFill(bgColor='FFFF00'))
        rule = Rule(type="expression", dxf=style)
        rule.formula = ['V2=1']
        sheet.conditional_formatting.add("V2:AD"+str((len(df) // mod) + 4), rule)

        for i in range((len(df) // mod) + 2):
            rule2 = Rule(type='expression', dxf=style)
            rule2.formula = ['AP' + str(5 + 13*i) + '=MAX($AP' + str(5 + 13*i) + ':$AW' + str(5 + 13*i) + ')']
            sheet.conditional_formatting.add('AP' + str(5 + 13*i) + ':AW' + str(12 + 13*i), rule2)

        file.save(os.path.join('output', filename))
#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))

